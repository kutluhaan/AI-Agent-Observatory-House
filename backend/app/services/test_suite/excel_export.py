"""
Excel export — F1.2

Bir TestRun'ı (özet + case sonuçları + assertion/judge detayları) .xlsx
çalışma kitabına dönüştürür. openpyxl ile stillenir.

Sayfalar:
  Özet           — run bilgisi + toplu metrikler
  Case Sonuçları — case başına özet (durum, assertion, token, maliyet, tutarlılık)
  Assertion'lar  — case × assertion satırları
  Judge'lar      — case × LLM-judge satırları
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="312E81")  # indigo-900
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_PASS_FILL = PatternFill("solid", fgColor="DCFCE7")    # green-100
_FAIL_FILL = PatternFill("solid", fgColor="FEE2E2")    # red-100
_ERR_FILL = PatternFill("solid", fgColor="FEF3C7")     # amber-100


def _status_fill(status: str) -> PatternFill | None:
    return {"passed": _PASS_FILL, "failed": _FAIL_FILL, "error": _ERR_FILL}.get(status)


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 2, max_width)


def build_workbook(run: Any, case_results: list, case_names: dict) -> bytes:
    wb = Workbook()

    # ── Özet ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Özet"
    s = run.summary or {}
    rows = [
        ("Test Run", str(run.id)),
        ("Durum", run.status),
        ("Başladı", run.started_at.isoformat() if run.started_at else "—"),
        ("Bitti", run.ended_at.isoformat() if run.ended_at else "—"),
        ("Toplam case", s.get("total")),
        ("Geçen", s.get("passed")),
        ("Kalan/Hata", (s.get("failed") or 0) + (s.get("error") or 0)),
        ("Geçme oranı", f"{round((s.get('pass_rate') or 0) * 100)}%"),
        ("Ort. gecikme (ms)", s.get("avg_latency_ms")),
        ("Toplam token", s.get("total_tokens")),
        ("Tahmini maliyet ($)", s.get("total_cost_usd")),
        ("Ort. judge skoru", s.get("avg_judge_score")),
    ]
    ws.append(["Metrik", "Değer"])
    _style_header(ws, 2)
    for k, v in rows:
        ws.append([k, v if v is not None else "—"])
    _autosize(ws)

    # ── Case Sonuçları ──────────────────────────────────────
    wc = wb.create_sheet("Case Sonuçları")
    wc.append([
        "Case", "Durum", "Assertion (geçen/toplam)", "Judge ort.",
        "Adım", "Token", "Maliyet ($)", "Gecikme (s)",
        "Tutarlılık", "Trace ID", "Çıktı (kısa)",
    ])
    _style_header(wc, 11)
    for r in case_results:
        ap = sum(1 for a in (r.assertions_results or []) if a.get("passed"))
        at = len(r.assertions_results or [])
        jscores = [j["score"] for j in (r.judge_results or []) if j.get("score") is not None]
        javg = round(sum(jscores) / len(jscores), 3) if jscores else "—"
        cons = r.consistency
        cons_str = f"{cons['passed_runs']}/{cons['runs']}" if cons else "—"
        out = (r.output or "").replace("\n", " ")[:200]
        row = [
            case_names.get(str(r.case_id), str(r.case_id)),
            r.status,
            f"{ap}/{at}" if at else "—",
            javg,
            r.steps_taken if r.steps_taken is not None else "—",
            r.total_tokens if r.total_tokens is not None else "—",
            round(r.cost_usd, 6) if r.cost_usd is not None else "—",
            round(r.latency_ms / 1000, 2) if r.latency_ms is not None else "—",
            cons_str,
            r.trace_id or "—",
            out,
        ]
        wc.append(row)
        fill = _status_fill(r.status)
        if fill:
            wc.cell(row=wc.max_row, column=2).fill = fill
    _autosize(wc)

    # ── Assertion'lar ───────────────────────────────────────
    wa = wb.create_sheet("Assertion'lar")
    wa.append(["Case", "Tip", "Geçti", "Beklenen", "Gerçek", "Mesaj"])
    _style_header(wa, 6)
    for r in case_results:
        for a in (r.assertions_results or []):
            wa.append([
                case_names.get(str(r.case_id), str(r.case_id)),
                a.get("type", ""),
                "✓" if a.get("passed") else "✗",
                str(a.get("expected", ""))[:300],
                str(a.get("actual", ""))[:300],
                str(a.get("message", ""))[:300],
            ])
            wa.cell(row=wa.max_row, column=3).fill = _PASS_FILL if a.get("passed") else _FAIL_FILL
    _autosize(wa)

    # ── Judge'lar ───────────────────────────────────────────
    wj = wb.create_sheet("Judge'lar")
    wj.append(["Case", "Metrik", "Skor", "Geçti", "Eşik", "Gerekçe / Hata"])
    _style_header(wj, 6)
    for r in case_results:
        for j in (r.judge_results or []):
            passed = j.get("passed")
            wj.append([
                case_names.get(str(r.case_id), str(r.case_id)),
                j.get("type", ""),
                j.get("score") if j.get("score") is not None else "—",
                "✓" if passed else ("—" if passed is None else "✗"),
                j.get("threshold"),
                str(j.get("rationale") or j.get("error") or "")[:400],
            ])
            if passed is True:
                wj.cell(row=wj.max_row, column=4).fill = _PASS_FILL
            elif passed is False:
                wj.cell(row=wj.max_row, column=4).fill = _FAIL_FILL
    _autosize(wj)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
