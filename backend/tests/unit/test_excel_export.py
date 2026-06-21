"""F1.2 — Excel export birim testleri (DB gerektirmez)."""
import io
import uuid
from datetime import UTC, datetime
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app.services.test_suite.excel_export import build_workbook


def _run():
    return SimpleNamespace(
        id=uuid.uuid4(),
        status="completed",
        started_at=datetime.now(UTC),
        ended_at=datetime.now(UTC),
        summary={
            "total": 2, "passed": 1, "failed": 1, "error": 0,
            "pass_rate": 0.5, "avg_latency_ms": 1200, "total_tokens": 3400,
            "total_cost_usd": 0.0021, "avg_judge_score": 0.82,
        },
    )


def _case(cid, status, **kw):
    return SimpleNamespace(
        case_id=cid, status=status,
        output=kw.get("output", "some output\nwith newline"),
        trace_id=kw.get("trace_id", "trace123"),
        latency_ms=kw.get("latency_ms", 1200),
        steps_taken=kw.get("steps_taken", 3),
        total_tokens=kw.get("total_tokens", 1700),
        cost_usd=kw.get("cost_usd", 0.001),
        assertions_results=kw.get("assertions_results", [
            {"type": "response_contains", "passed": True, "expected": "x", "actual": "x", "message": "OK"},
        ]),
        judge_results=kw.get("judge_results", [
            {"type": "task_completion", "score": 0.9, "passed": True, "threshold": 0.7, "rationale": "good"},
        ]),
        consistency=kw.get("consistency"),
    )


def test_build_workbook_produces_valid_xlsx():
    c1, c2 = uuid.uuid4(), uuid.uuid4()
    run = _run()
    results = [
        _case(c1, "passed"),
        _case(c2, "failed", consistency={"runs": 3, "passed_runs": 1}),
    ]
    names = {str(c1): "case-one", str(c2): "case-two"}

    data = build_workbook(run, results, names)
    assert isinstance(data, (bytes, bytearray)) and len(data) > 0

    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Özet", "Case Sonuçları", "Assertion'lar", "Judge'lar"]

    # Özet'te geçme oranı %50
    summary = wb["Özet"]
    vals = {row[0].value: row[1].value for row in summary.iter_rows(min_row=2)}
    assert vals["Geçme oranı"] == "50%"
    assert vals["Toplam token"] == 3400

    # Case sayfasında 2 satır (başlık hariç)
    cases = wb["Case Sonuçları"]
    assert cases.max_row == 3  # header + 2

    # Case adları yazıldı
    case_names_in_sheet = {cases.cell(row=r, column=1).value for r in (2, 3)}
    assert case_names_in_sheet == {"case-one", "case-two"}


def test_build_workbook_handles_empty_results():
    data = build_workbook(_run(), [], {})
    wb = load_workbook(io.BytesIO(data))
    assert "Özet" in wb.sheetnames
